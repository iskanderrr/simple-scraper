import requests as rq
from bs4 import BeautifulSoup
import openpyxl
import string
URL = "" #url removed
page = rq.get(URL)

soup = BeautifulSoup(page.content, "html.parser")
results = soup.find(id="tokoo-shop-view-content") #products container

products_list = results.find_all("div", class_="product-outer") #products grid elements

workbook = openpyxl.Workbook()
sheet = workbook.active

titles = ["id","name","brand","price",'description'] #titles for the sheet
for i in range(5):
    sheet.cell(1, i+1).value=titles[i]
i=2
for product in products_list:
    try:
        sheet.cell(i, 1).value=i                                                                          #id
        sheet.cell(i, 2).value = product.find("h2", class_="woocommerce-loop-product__title").text.strip()#product name
        sheet.cell(i, 3).value =product.find("div", class_="pwb-brands-in-loop").text.strip()             #brand
        sheet.cell(i, 4).value =product.find("bdi").text.strip()                                          #price
        sheet.cell(i, 5).value = product.find("p").text.strip()                                           #description

        im=rq.get(product.find("img")["src"]).content #get image data
        #the file name is the id.jpg
        with open("images/"+str(i)+".jpg", 'wb') as handler:
            handler.write(im)

        i+=1

    except:
        pass
workbook.save('output.xlsx') #save the file
workbook.close()
